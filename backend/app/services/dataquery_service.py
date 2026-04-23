"""
智能问数核心服务 (NL2SQL)

流程：
1. 获取数据源的表元数据（schema 信息）
2. 构造 System Prompt + 用户问题
3. 调用 LLM 生成 SQL
4. SQL 安全校验（仅允许 SELECT）
5. 执行 SQL 并获取结果
6. AI 生成自然语言摘要 + 图表推荐
7. 保存查询历史
"""

import json
import logging
import re
import time
from typing import AsyncGenerator, List, Optional

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.ai import call_llm
from app.crud.dataquery import datasource as ds_crud
from app.crud.dataquery import history as history_crud
from app.services import db_connector

logger = logging.getLogger(__name__)

# ============ SQL 安全校验 ============

# 禁止的 SQL 关键字（大写）
_FORBIDDEN_KEYWORDS = [
    "DROP", "TRUNCATE", "ALTER", "CREATE", "DELETE", "UPDATE", "INSERT",
    "REPLACE", "GRANT", "REVOKE", "RENAME", "LOCK", "UNLOCK", "FLUSH",
    "KILL", "LOAD", "CALL", "EXEC", "EXECUTE",
]

_MAX_RESULT_ROWS = 1000


def validate_sql(sql: str) -> tuple[bool, str]:
    """
    校验 SQL 安全性，仅允许 SELECT 查询。

    Returns:
        (is_safe, error_message)
    """
    if not sql or not sql.strip():
        return False, "SQL 不能为空"

    cleaned = sql.strip().rstrip(";")

    # 检查是否以 SELECT/WITH/EXPLAIN 开头
    first_word = cleaned.split()[0].upper() if cleaned.split() else ""
    if first_word not in ("SELECT", "WITH", "EXPLAIN", "SHOW", "DESCRIBE", "DESC"):
        return False, f"仅允许查询语句，不允许 {first_word} 操作"

    # 检查禁止关键字（排除在字符串引号内的）
    upper_sql = cleaned.upper()
    for kw in _FORBIDDEN_KEYWORDS:
        # 使用词边界匹配
        pattern = rf"\b{kw}\b"
        if re.search(pattern, upper_sql):
            return False, f"SQL 中包含禁止的关键字: {kw}"

    # 检查是否有多条语句（分号分隔）
    # 简化检查：去除字符串中的分号后再判断
    if ";" in cleaned:
        return False, "不允许执行多条 SQL 语句"

    return True, ""


def ensure_limit(sql: str, max_rows: int = _MAX_RESULT_ROWS) -> str:
    """确保 SQL 有 LIMIT 限制（仅对 SELECT/WITH 语句生效）"""
    stripped = sql.strip()
    first_word = stripped.split()[0].upper() if stripped.split() else ""
    # SHOW / DESCRIBE / DESC / EXPLAIN 不支持 LIMIT
    if first_word not in ("SELECT", "WITH"):
        return stripped.rstrip(";")
    upper_sql = stripped.upper()
    if "LIMIT" not in upper_sql:
        stripped = stripped.rstrip().rstrip(";")
        stripped += f" LIMIT {max_rows}"
    return stripped


# ============ Schema Prompt 构建 ============


def build_schema_prompt(tables_meta: list, db_type: str = "mysql") -> str:
    """
    根据表元数据构建 Schema 描述，用于 LLM 理解数据库结构。
    """
    if not tables_meta:
        return "（数据库中没有表信息）"

    dialect = "MySQL" if db_type.lower() == "mysql" else "PostgreSQL"
    lines = [f"数据库类型: {dialect}", "数据库表结构如下：", ""]

    for table in tables_meta:
        table_name = table.get("table_name", "")
        table_comment = table.get("table_comment", "")
        custom_desc = table.get("custom_description", "")
        columns = table.get("columns", [])

        header = f"### 表: `{table_name}`"
        if table_comment:
            header += f" ({table_comment})"
        lines.append(header)

        if custom_desc:
            lines.append(f"业务说明: {custom_desc}")

        lines.append("| 字段 | 类型 | 说明 | 主键 | 可空 |")
        lines.append("|------|------|------|------|------|")
        for col in columns:
            pk = "是" if col.get("is_pk") else ""
            nullable = "是" if col.get("nullable") else "否"
            comment = col.get("comment", "")
            lines.append(
                f"| `{col['name']}` | {col['type']} | {comment} | {pk} | {nullable} |"
            )

        # 样例数据
        sample_data = table.get("sample_data")
        if sample_data:
            lines.append(f"\n样例数据（前 {len(sample_data)} 行）:")
            lines.append("```json")
            lines.append(json.dumps(sample_data, ensure_ascii=False, indent=2)[:500])
            lines.append("```")

        lines.append("")

    return "\n".join(lines)


# ============ NL2SQL System Prompt ============

NL2SQL_SYSTEM_PROMPT = """你是一个专业的数据分析 SQL 助手。用户会用自然语言描述他们想要查询的数据，你需要根据提供的数据库表结构生成正确的 SQL 查询。

## 工作规范

1. **只生成 SELECT 查询**，绝对不要生成 INSERT/UPDATE/DELETE/DROP 等写操作
2. 根据用户问题和表结构生成准确的 SQL
3. SQL 必须兼容目标数据库方言
4. 当结果可能很多时，自动添加合理的 LIMIT
5. 使用清晰的表/字段别名提升可读性
6. 对于统计类问题，优先使用聚合函数（COUNT, SUM, AVG 等）
7. 对于时间相关的问题，使用合适的日期函数

## 输出格式

你必须严格按照以下 JSON 格式输出，不要有任何多余的文字：

```json
{
  "sql": "你生成的 SQL 查询语句",
  "explanation": "简要说明这个 SQL 做了什么",
  "chart_recommendation": {
    "type": "table|bar|line|pie|none",
    "x_field": "X轴字段名（图表时填写）",
    "y_field": "Y轴字段名（图表时填写）",
    "reason": "为什么推荐这种图表"
  }
}
```

## 图表推荐规则
- 数据行数 <= 1 或纯文本查询 → type: "table"
- 有时间/日期列 + 数值列 → type: "line"
- 有分类列 + 数值列（分类 <= 20）→ type: "bar"
- 有分类列 + 占比/比例值 → type: "pie"
- 其他情况 → type: "table"
"""


SUMMARY_SYSTEM_PROMPT = """你是一个数据分析师。根据用户的问题和 SQL 查询结果，用简洁的自然语言总结分析结果。

## 规范
1. 直接回答用户的问题
2. 提炼关键数据指标
3. 如果数据中有明显的趋势或异常，指出来
4. 回答要简洁，控制在 2-4 句话
5. 使用数字和百分比让回答更精确
"""


# ============ 核心服务函数 ============


async def ask(
    db: AsyncSession,
    user_id: int,
    datasource_id: int,
    question: str,
    conversation_history: Optional[List[dict]] = None,
) -> dict:
    """
    智能问数主流程（非流式）

    Returns:
        {
            "history_id": int,
            "question": str,
            "generated_sql": str,
            "result_data": list[dict],
            "result_summary": str,
            "row_count": int,
            "execution_time": float,
            "chart_type": str,
            "chart_config": dict,
            "status": "success" | "error",
            "error_message": str | None,
        }
    """
    # 1. 获取数据源
    datasource = await ds_crud.get_datasource(db, datasource_id)
    if not datasource:
        return {"status": "error", "error_message": "数据源不存在"}

    if datasource.status != "active":
        return {"status": "error", "error_message": "数据源已禁用"}

    # 2. 获取表元数据
    tables_meta_db = await ds_crud.get_table_metadata_list(db, datasource_id)
    tables_meta = []
    for tm in tables_meta_db:
        tables_meta.append({
            "table_name": tm.table_name,
            "table_comment": tm.table_comment,
            "columns": tm.columns or [],
            "custom_description": tm.custom_description,
            "sample_data": tm.sample_data,
        })

    if not tables_meta:
        return {
            "status": "error",
            "error_message": "数据源尚未同步表结构，请先在数据源管理中同步元数据",
        }

    # 3. 构造 LLM 消息
    schema_text = build_schema_prompt(tables_meta, datasource.db_type)
    messages = [
        {"role": "system", "content": NL2SQL_SYSTEM_PROMPT},
        {
            "role": "user",
            "content": f"{schema_text}\n\n---\n\n用户问题：{question}",
        },
    ]

    # 加入对话历史（最后 6 条）
    if conversation_history:
        for msg in conversation_history[-6:]:
            messages.insert(-1, msg)

    # 4. 调用 LLM 生成 SQL
    generated_sql = ""
    explanation = ""
    chart_type = "table"
    chart_config = None
    error_message = None
    result_data = None
    result_summary = None
    row_count = 0
    execution_time = None
    status = "success"

    try:
        response = await call_llm(messages, stream=False)
        content = response.choices[0].message.content or ""

        # 解析 LLM 返回的 JSON
        parsed = _parse_llm_json(content)
        generated_sql = parsed.get("sql", "").strip()
        explanation = parsed.get("explanation", "")
        chart_rec = parsed.get("chart_recommendation", {})
        chart_type = chart_rec.get("type", "table") if chart_rec else "table"
        chart_config = chart_rec if chart_rec else None

        if not generated_sql:
            status = "error"
            error_message = "AI 未能生成有效的 SQL"

    except Exception as e:
        logger.error(f"[NL2SQL] LLM 调用失败: {e}", exc_info=True)
        status = "error"
        error_message = f"AI 服务调用失败: {str(e)}"

    # 5. SQL 安全校验
    if status == "success" and generated_sql:
        is_safe, err = validate_sql(generated_sql)
        if not is_safe:
            status = "error"
            error_message = f"SQL 安全校验失败: {err}"

    # 6. 执行 SQL
    if status == "success" and generated_sql:
        safe_sql = ensure_limit(generated_sql)
        try:
            query_result = await db_connector.execute_query(
                datasource, safe_sql, timeout=30
            )
            if "error" in query_result and query_result["error"]:
                status = "error"
                error_message = f"SQL 执行失败: {query_result['error']}"
            else:
                result_data = query_result["rows"]
                row_count = query_result["row_count"]
                execution_time = query_result["execution_time"]
        except Exception as e:
            status = "error"
            error_message = f"SQL 执行异常: {str(e)}"

    # 7. 生成自然语言摘要
    if status == "success" and result_data is not None:
        try:
            result_summary = await _generate_summary(
                question, generated_sql, result_data, row_count
            )
        except Exception as e:
            logger.warning(f"[NL2SQL] 生成摘要失败: {e}")
            result_summary = f"查询返回 {row_count} 条数据。"

        # 智能调整图表类型
        if row_count <= 1:
            chart_type = "table"
        elif row_count > 50:
            chart_type = chart_type or "table"

    # 8. 保存查询历史
    history = await history_crud.create_query_history(
        db=db,
        datasource_id=datasource_id,
        user_id=user_id,
        question=question,
        generated_sql=generated_sql,
        result_summary=result_summary,
        result_data=result_data[:200] if result_data else None,  # 最多保存 200 行
        row_count=row_count,
        execution_time=execution_time,
        status=status,
        error_message=error_message,
        chart_type=chart_type,
        chart_config=chart_config,
    )

    return {
        "history_id": history.id,
        "question": question,
        "generated_sql": generated_sql,
        "explanation": explanation,
        "result_data": result_data,
        "result_summary": result_summary,
        "row_count": row_count,
        "execution_time": execution_time,
        "chart_type": chart_type,
        "chart_config": chart_config,
        "status": status,
        "error_message": error_message,
    }


async def ask_stream(
    db: AsyncSession,
    user_id: int,
    datasource_id: int,
    question: str,
    conversation_history: Optional[List[dict]] = None,
) -> AsyncGenerator[str, None]:
    """
    智能问数流式输出（SSE）

    SSE 事件格式：
    - data: [SQL]{json}           — 生成的 SQL
    - data: [EXECUTING]           — 正在执行 SQL
    - data: [RESULT]{json}        — 查询结果
    - data: 普通文本              — 摘要流式输出
    - data: [DONE]{json}          — 完成，包含 history_id
    - data: [ERROR]msg            — 错误
    """

    def _sse(text: str) -> str:
        return "data: " + text.replace("\n", "\\n") + "\n\n"

    # 1. 获取数据源
    datasource = await ds_crud.get_datasource(db, datasource_id)
    if not datasource:
        yield "data: [ERROR]数据源不存在\n\n"
        return
    if datasource.status != "active":
        yield "data: [ERROR]数据源已禁用\n\n"
        return

    # 2. 获取表元数据
    tables_meta_db = await ds_crud.get_table_metadata_list(db, datasource_id)
    tables_meta = []
    for tm in tables_meta_db:
        tables_meta.append({
            "table_name": tm.table_name,
            "table_comment": tm.table_comment,
            "columns": tm.columns or [],
            "custom_description": tm.custom_description,
            "sample_data": tm.sample_data,
        })

    if not tables_meta:
        yield "data: [ERROR]数据源尚未同步表结构，请先同步元数据\n\n"
        return

    # 3. 调用 LLM 生成 SQL
    schema_text = build_schema_prompt(tables_meta, datasource.db_type)
    messages = [
        {"role": "system", "content": NL2SQL_SYSTEM_PROMPT},
        {"role": "user", "content": f"{schema_text}\n\n---\n\n用户问题：{question}"},
    ]
    if conversation_history:
        for msg in conversation_history[-6:]:
            messages.insert(-1, msg)

    generated_sql = ""
    explanation = ""
    chart_type = "table"
    chart_config = None

    try:
        response = await call_llm(messages, stream=False)
        content = response.choices[0].message.content or ""
        parsed = _parse_llm_json(content)
        generated_sql = parsed.get("sql", "").strip()
        explanation = parsed.get("explanation", "")
        chart_rec = parsed.get("chart_recommendation", {})
        chart_type = chart_rec.get("type", "table") if chart_rec else "table"
        chart_config = chart_rec if chart_rec else None

        if not generated_sql:
            yield "data: [ERROR]AI 未能生成有效的 SQL\n\n"
            return
    except Exception as e:
        yield f"data: [ERROR]AI 服务调用失败: {str(e)}\n\n"
        return

    # 4. 发送 SQL
    sql_data = json.dumps(
        {"sql": generated_sql, "explanation": explanation},
        ensure_ascii=False,
    )
    yield f"data: [SQL]{sql_data}\n\n"

    # 5. SQL 安全校验
    is_safe, err = validate_sql(generated_sql)
    if not is_safe:
        yield f"data: [ERROR]SQL 安全校验失败: {err}\n\n"
        return

    # 6. 执行 SQL
    yield "data: [EXECUTING]\n\n"
    safe_sql = ensure_limit(generated_sql)
    try:
        query_result = await db_connector.execute_query(datasource, safe_sql, timeout=30)
        if "error" in query_result and query_result["error"]:
            # 保存失败记录
            await history_crud.create_query_history(
                db=db, datasource_id=datasource_id, user_id=user_id,
                question=question, generated_sql=generated_sql,
                status="error", error_message=query_result["error"],
            )
            yield f"data: [ERROR]SQL 执行失败: {query_result['error']}\n\n"
            return
    except Exception as e:
        yield f"data: [ERROR]SQL 执行异常: {str(e)}\n\n"
        return

    result_data = query_result["rows"]
    row_count = query_result["row_count"]
    execution_time = query_result["execution_time"]

    # 智能调整图表类型
    if row_count <= 1:
        chart_type = "table"

    # 7. 发送结果
    # 前端展示最多 200 行，超出的提醒导出
    display_data = result_data[:200]
    result_payload = json.dumps(
        {
            "columns": query_result["columns"],
            "rows": display_data,
            "row_count": row_count,
            "execution_time": execution_time,
            "chart_type": chart_type,
            "chart_config": chart_config,
            "has_more": row_count > 200,
        },
        ensure_ascii=False,
        default=str,
    )
    yield f"data: [RESULT]{result_payload}\n\n"

    # 8. 流式生成摘要
    result_summary = ""
    try:
        summary_messages = [
            {"role": "system", "content": SUMMARY_SYSTEM_PROMPT},
            {
                "role": "user",
                "content": (
                    f"用户问题：{question}\n\n"
                    f"执行的 SQL：{generated_sql}\n\n"
                    f"查询结果（共 {row_count} 行）：\n"
                    f"{json.dumps(display_data[:20], ensure_ascii=False, default=str)}"
                ),
            },
        ]
        stream_resp = await call_llm(summary_messages, stream=True)
        async for chunk in stream_resp:
            delta = chunk.choices[0].delta
            if delta.content:
                result_summary += delta.content
                yield _sse(delta.content)
    except Exception as e:
        result_summary = f"查询返回 {row_count} 条数据。"
        yield _sse(result_summary)

    # 9. 保存查询历史
    history = await history_crud.create_query_history(
        db=db,
        datasource_id=datasource_id,
        user_id=user_id,
        question=question,
        generated_sql=generated_sql,
        result_summary=result_summary,
        result_data=display_data,
        row_count=row_count,
        execution_time=execution_time,
        status="success",
        chart_type=chart_type,
        chart_config=chart_config,
    )

    # 10. 完成
    done_data = json.dumps({"history_id": history.id}, ensure_ascii=False)
    yield f"data: [DONE]{done_data}\n\n"


# ============ 内部工具函数 ============


def _parse_llm_json(content: str) -> dict:
    """解析 LLM 返回的 JSON（兼容 ```json 包裹）"""
    content = content.strip()

    # 去除 markdown 代码块
    if content.startswith("```"):
        lines = content.split("\n")
        # 去除首行和末行
        if lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        content = "\n".join(lines).strip()

    try:
        return json.loads(content)
    except json.JSONDecodeError:
        # 尝试从文本中提取 JSON
        match = re.search(r"\{[\s\S]*\}", content)
        if match:
            try:
                return json.loads(match.group())
            except json.JSONDecodeError:
                pass
        # 如果完全无法解析，尝试提取 SQL
        sql_match = re.search(r"(?:SELECT|WITH)\s+.+", content, re.IGNORECASE | re.DOTALL)
        if sql_match:
            return {"sql": sql_match.group().strip(), "explanation": ""}
        return {}


async def _generate_summary(
    question: str,
    sql: str,
    result_data: list,
    row_count: int,
) -> str:
    """生成查询结果的自然语言摘要"""
    preview = result_data[:20] if result_data else []
    messages = [
        {"role": "system", "content": SUMMARY_SYSTEM_PROMPT},
        {
            "role": "user",
            "content": (
                f"用户问题：{question}\n\n"
                f"执行的 SQL：{sql}\n\n"
                f"查询结果（共 {row_count} 行）：\n"
                f"{json.dumps(preview, ensure_ascii=False, default=str)}"
            ),
        },
    ]

    response = await call_llm(messages, stream=False)
    return response.choices[0].message.content or f"查询返回 {row_count} 条数据。"


async def export_to_excel(
    db: AsyncSession,
    history_id: int,
    datasource_id: int,
) -> Optional[bytes]:
    """将查询结果导出为 Excel 文件"""
    try:
        from openpyxl import Workbook
        from io import BytesIO
    except ImportError:
        raise RuntimeError("需要安装 openpyxl: pip install openpyxl")

    history = await history_crud.get_query_history(db, history_id)
    if not history or not history.generated_sql:
        return None

    # 如果历史记录中有结果数据且行数不多，直接用
    if history.result_data and history.row_count <= 200:
        data = history.result_data
    else:
        # 重新执行查询获取完整数据
        datasource = await ds_crud.get_datasource(db, datasource_id)
        if not datasource:
            return None
        safe_sql = ensure_limit(history.generated_sql, max_rows=50000)
        result = await db_connector.execute_query(datasource, safe_sql, timeout=60)
        if "error" in result and result["error"]:
            return None
        data = result["rows"]

    if not data:
        return None

    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "查询结果"

    # 写入表头
    headers = list(data[0].keys()) if data else []
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 写入数据
    for row_idx, row in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row.get(header)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # 保存到 BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
